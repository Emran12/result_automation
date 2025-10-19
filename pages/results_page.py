import time # Import time for simple waits (should be replaced by WebDriverWait for production)
import os
import re
import cv2
import re
import numpy as np
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class ResultsPage:
    """
    Page Object Model for the Result Form page.
    Contains all locators and methods for interacting with the form elements.
    """
    
    def __init__(self, driver):
        """Initializes the page object with the WebDriver instance."""
        self.driver = driver
        
        # --- Locators (Defined using By tuples) ---
        # STANDARD DROPDOWN LOCATORS (Assumed to be <select> elements)
        self.EXAM_DROPDOWN = (By.ID, "exam")        
        self.YEAR_DROPDOWN = (By.ID, "year")        
        self.RESULT_TYPE_DROPDOWN = (By.ID, "result_type") 

        # CUSTOM DROPDOWN LOCATORS (For the complex Board field)
        # 1. Locator for the visible element (the button/div that opens the list)
        self.BOARD_OPEN_BUTTON = (By.ID, "board") 
        
        
        # INPUT AND BUTTON LOCATORS
        self.ROLL_INPUT = (By.ID, "roll")      
        self.REG_INPUT = (By.XPATH, '//div[@id="row_reg"]//input[@id="reg"]')        
        self.CAPTCHA_IMAGE = (By.ID, "captcha_img") 
        self.CAPTCHA_INPUT = (By.ID, "captcha") 
        self.SUBMIT_BUTTON = (By.ID, "submit") 

    # --- Interaction Methods ---
    def select_standard_dropdown(self, button, data, key):
        """
        Handles selection for a custom-styled dropdown menu.
        Clicks the visible dropdown trigger, then selects the desired option by text.
        """
        try:
            print(f"Selecting {key}: {data[key]} using custom click sequence.")

            # Step 1: Click dropdown to open the visible list
            open_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(button)
            )
            open_button.click()
            time.sleep(0.5)

            # Step 2: Find the visible option (adjust locator depending on HTML structure)
            option_xpath = f'//select[@id="{key}"]/option[@value="{data[key]}"]'
            print("option_xpath:", option_xpath)

            option_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, option_xpath))
            )

            # Step 3: Scroll to and click the visible option
            self.driver.execute_script(
                "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", 
                option_element
            )
            time.sleep(0.3)
            option_element.click()

            # Step 4: Click outside to close dropdown
            self.driver.find_element(By.TAG_NAME, "body").click()
            print(f"✅ Successfully selected custom {key}: {data[key]}.")

        except Exception as e:
            print(f"❌ Error selecting custom board dropdown value: {e}")



    def fill_input_field(self, locator, text):
        """Generic method to enter text into an input field."""
        self.driver.find_element(*locator).send_keys(text)
        print(f"Entered text in {locator[1]}: {text}")
        
    def submit_form(self):
        """Clicks the main submit button."""
        self.driver.find_element(*self.SUBMIT_BUTTON).click()
        print("Form submitted.")

    # --- Specific Form Interaction Wrappers ---
    
    def fill_form_data(self, data):
        """Fills all fields using the data dictionary."""
        
        # Select Dropdowns
        # Board, EXAM, YEAR, RESULT_TYPE use the standard <select> method
        self.select_standard_dropdown(self.BOARD_OPEN_BUTTON, data, "board") 
        self.select_standard_dropdown(self.EXAM_DROPDOWN, data, "exam")
        self.select_standard_dropdown(self.YEAR_DROPDOWN, data, "year")
        self.select_standard_dropdown(self.RESULT_TYPE_DROPDOWN, data, "result_type")
       
        # Fill Inputs
        self.fill_input_field(self.ROLL_INPUT, data["roll_number"])
        self.fill_input_field(self.REG_INPUT, data["reg_number"])
        time.sleep(10)

    def enter_captcha_solution(self, solution):
        """Enters the provided CAPTCHA solution."""
        self.fill_input_field(self.CAPTCHA_INPUT, solution)
        
    def save_captcha_image(self):
        """Saves the CAPTCHA image for manual or service solving."""
        try:
            captcha_element = self.driver.find_element(*self.CAPTCHA_IMAGE)
            captcha_element.screenshot("current_captcha.png")
            # number = self.extract_number_from_image("current_captcha.png")
            # print("number : ", number)
            # self.enter_captcha_solution(number)
            return True
        except Exception as e:
            print(f"Could not save CAPTCHA image: {e}")
            return False
    
    def extract_number_from_image(self, image_path):
        """
        Extracts numbers from a CAPTCHA or image using Tesseract OCR with preprocessing.
        """

        try:
            img = cv2.imread(image_path)
    
            # 1. Define source points (top-left, top-right, bottom-right, bottom-left) 
            #    around the curved number in the original image.
            #    ***These coordinates must be manually determined for your image***
            #    Example points for a typical curved banner:
            src_points = np.float32([
                [50, 100],  # Top-Left (approx)
                [400, 100], # Top-Right (approx)
                [450, 250], # Bottom-Right (approx)
                [0, 250]    # Bottom-Left (approx)
            ])

            # 2. Define destination points (a flat rectangle)
            width, height = 400, 150
            dst_points = np.float32([
                [0, 0], 
                [width, 0], 
                [width, height], 
                [0, height]
            ])

            # 3. Get the perspective transformation matrix
            matrix = cv2.getPerspectiveTransform(src_points, dst_points)

            # 4. Perform the perspective warp
            unwarped_img = cv2.warpPerspective(img, matrix, (width, height))

            # 5. Convert the unwarped OpenCV image to a PIL Image for Tesseract
            pil_img = Image.fromarray(cv2.cvtColor(unwarped_img, cv2.COLOR_BGR2RGB))
            
            # --- Continue with OCR on the unwarped image ---
            pil_img = pil_img.convert('L') # Grayscale
            # You might still need to apply binarization here
            
            config = r'--oem 3 --psm 7 -c tessedit_char_whitelist=0123456789'
            text = pytesseract.image_to_string(pil_img, config=config)
            print("text: ", text)
            return "".join(filter(str.isdigit, text))

        except pytesseract.TesseractNotFoundError:
            return "Tesseract is not installed or not in your PATH. Please install Tesseract OCR."
        except FileNotFoundError:
            return f"Error: Image file not found at {image_path}"
        except Exception as e:
            # This catches issues like PIL not being able to open the file type
            return f"An unexpected error occurred: {e}"

       

    def take_result_data(self, data):
        # --- Find all <tbody> tags ---
        all_tbody = self.driver.find_elements(By.TAG_NAME, "tbody")

        if len(all_tbody) < 2:
            print("⚠️ Not enough tbody elements found.")
            return

        student_tbody = all_tbody[0]
        subject_tbody = all_tbody[1]

        # --- Extract student info ---
        student_name = group = gender = gpa = "N/A"

        rows = student_tbody.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            texts = [c.text.strip() for c in cells]

            if not texts:
                continue

            if "Name of Student" in texts[0]:
                student_name = texts[1]
            elif "Group" in texts[0]:
                group = texts[1]
                for t in texts:
                    if "Gender:" in t:
                        gender = t.split(":")[-1].strip()
            elif "Result" in texts[0]:
                for t in texts:
                    if "GPA=" in t:
                        gpa = t.split("=")[-1].strip()

        print("Student Name:", student_name)
        print("Group:", group)
        print("Gender:", gender)
        print("GPA:", gpa)

        # --- Extract subject info ---
        subject_data = []
        subject_rows = subject_tbody.find_elements(By.TAG_NAME, "tr")
        for row in subject_rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) == 4:
                subject_code = cols[0].text.strip()
                subject_name = cols[1].text.strip()
                marks = cols[2].text.strip()
                grade = cols[3].text.strip()
                subject_data.append((subject_code, subject_name, marks, grade))

        # --- Save both student info & subjects to Excel ---
        self.save_to_excel(data['roll_number'], student_name, group, gender, gpa, subject_data)


    def save_to_excel(self, roll_number, name, group, gender, gpa, subject_data):

        file_path = "results.xlsx"

        # If file doesn't exist → create workbook with headers
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            # Base headers
            headers = ["Roll Number", "Student Name", "Group", "Gender", "GPA"]

            # Add headers for subjects dynamically (we’ll support up to 20 subjects)
            for i in range(1, 9):
                headers += [f"Subject {i} Code", f"Subject {i} Name", f"Marks {i}", f"Grade {i}"]

            ws.append(headers)
            wb.save(file_path)

        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # Prepare row data
        row = [roll_number, name, group, gender, gpa]
        for sub in subject_data:
            row += [sub[0], sub[1], sub[2], sub[3]]

        # Fill remaining subject columns if less than 20 subjects
        missing = (8 - len(subject_data)) * 4
        row += [""] * missing

        ws.append(row)
        wb.save(file_path)
        print(f"✅ Saved result for {name} in a single row.")

